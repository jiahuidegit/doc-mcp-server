"""
Excel 文档分析器
支持复杂Excel(合并单元格、多层表头等)
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Tuple, Optional
import os

from .base import (
    BaseAnalyzer,
    DocumentMeta,
    DocumentFormat,
    SectionInfo,
    FieldInfo
)


class ExcelAnalyzer(BaseAnalyzer):
    """Excel文档分析器"""

    def __init__(self, file_path: str, sheet_name: Optional[str] = None):
        super().__init__(file_path)
        self.sheet_name = sheet_name
        self.wb = None
        self.ws = None
        self.merged_cells_map = {}
        self._load_workbook()

    def _load_workbook(self):
        """加载Excel工作簿"""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"文件不存在: {self.file_path}")

        self.wb = openpyxl.load_workbook(self.file_path, data_only=False)

        # 选择工作表
        if self.sheet_name:
            if self.sheet_name not in self.wb.sheetnames:
                raise ValueError(f"工作表不存在: {self.sheet_name}")
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.active

        # 构建合并单元格映射
        self._build_merged_cells_map()

    def _build_merged_cells_map(self):
        """构建合并单元格映射表"""
        self.merged_cells_map = {}
        for merged_range in self.ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row, merged_range.min_col
            max_row, max_col = merged_range.max_row, merged_range.max_col

            # 所有合并单元格都指向左上角的主单元格
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    self.merged_cells_map[(row, col)] = (min_row, min_col, max_row, max_col)

    def _get_cell_value(self, row: int, col: int) -> Any:
        """获取单元格值(处理合并单元格)"""
        if (row, col) in self.merged_cells_map:
            main_row, main_col, _, _ = self.merged_cells_map[(row, col)]
            value = self.ws.cell(main_row, main_col).value
        else:
            value = self.ws.cell(row, col).value

        return str(value).strip() if value else ""

    def _get_cell_coordinate(self, row: int, col: int) -> str:
        """获取单元格坐标(Excel格式)"""
        return f"{get_column_letter(col)}{row}"

    def _detect_sections(self) -> List[SectionInfo]:
        """检测章节"""
        sections = []
        current_section = None

        for row in range(1, self.ws.max_row + 1):
            first_col_value = self._get_cell_value(row, 1)

            # 检测章节标题(包含"部分"、"信息概要"等关键词)
            is_section_title = any(keyword in first_col_value for keyword in
                                   ['部分', '信息概要', '明细', '附件', '记录'])

            if is_section_title and len(first_col_value) > 3:
                # 结束上一个章节
                if current_section:
                    current_section.end_row = row - 1
                    sections.append(current_section)

                # 开始新章节
                current_section = SectionInfo(
                    title=first_col_value,
                    start_row=row,
                    end_row=self.ws.max_row,  # 临时值
                    start_col=1,
                    end_col=self.ws.max_column
                )

        # 添加最后一个章节
        if current_section:
            sections.append(current_section)

        return sections

    def _extract_fields(self, sections: List[SectionInfo]) -> List[FieldInfo]:
        """提取字段信息"""
        fields = []

        for section in sections:
            # 提取每个章节的前几行作为表头区域
            header_rows = min(5, section.end_row - section.start_row + 1)

            for row in range(section.start_row, section.start_row + header_rows):
                for col in range(1, min(self.ws.max_column + 1, 25)):  # 最多25列
                    value = self._get_cell_value(row, col)

                    # 跳过空值和过长的值(可能不是字段名)
                    if not value or len(value) > 50:
                        continue

                    # 生成字段键
                    field_key = f"{section.title}_{value}"
                    coord = self._get_cell_coordinate(row, col)

                    field = FieldInfo(
                        key=field_key,
                        name=value,
                        coord=coord,
                        row=row,
                        col=col,
                        section=section.title
                    )
                    fields.append(field)

        return fields

    def analyze(self) -> Dict[str, Any]:
        """分析Excel结构"""
        # 获取文档元数据
        file_size = os.path.getsize(self.file_path)
        meta = DocumentMeta(
            format=DocumentFormat.EXCEL,
            file_path=self.file_path,
            file_size=file_size,
            page_count=len(self.wb.sheetnames)
        )

        # 检测章节
        sections = self._detect_sections()

        # 提取字段
        fields = self._extract_fields(sections)
        meta.total_fields = len(fields)

        # 生成摘要
        summary = self._generate_summary(meta, sections, fields)

        return {
            'meta': meta,
            'sections': sections,
            'fields': fields,
            'summary': summary,
            'merged_cells_count': len(self.merged_cells_map)
        }

    def _generate_summary(
        self,
        meta: DocumentMeta,
        sections: List[SectionInfo],
        fields: List[FieldInfo]
    ) -> str:
        """生成结构摘要"""
        summary = f"""
Excel文档结构摘要
================
工作表: {self.ws.title}
总行数: {self.ws.max_row}
总列数: {self.ws.max_column}
章节数: {len(sections)}
字段数: {len(fields)}
合并单元格: {len(self.merged_cells_map)}

章节列表:
"""
        for i, section in enumerate(sections, 1):
            row_range = f"{section.start_row}-{section.end_row}"
            summary += f"  {i}. {section.title} (第{row_range}行)\n"

        return summary

    def get_field_value(self, field_key: str) -> Any:
        """获取字段值"""
        structure = self.get_structure()

        # 查找字段
        field = next((f for f in structure['fields'] if f.key == field_key), None)
        if not field:
            raise ValueError(f"字段不存在: {field_key}")

        return self._get_cell_value(field.row, field.col)

    def set_field_value(self, field_key: str, value: Any):
        """设置字段值"""
        structure = self.get_structure()

        # 查找字段
        field = next((f for f in structure['fields'] if f.key == field_key), None)
        if not field:
            raise ValueError(f"字段不存在: {field_key}")

        # 写入值
        self.ws.cell(field.row, field.col, value)

    def get_section_data(self, section_name: str) -> Dict[str, Any]:
        """获取章节数据"""
        structure = self.get_structure()

        # 查找章节的所有字段
        section_fields = [f for f in structure['fields'] if f.section == section_name]

        data = {}
        for field in section_fields:
            # 提取字段名(去掉章节前缀)
            field_name = field.name
            value = self._get_cell_value(field.row, field.col)
            data[field_name] = value

        return data

    def save(self, output_path: Optional[str] = None):
        """保存Excel文件"""
        save_path = output_path or self.file_path
        self.wb.save(save_path)

    def list_sections(self) -> List[str]:
        """列出所有章节名"""
        structure = self.get_structure()
        return [section.title for section in structure['sections']]

    def list_fields(self, section_name: Optional[str] = None) -> List[str]:
        """列出字段"""
        structure = self.get_structure()

        if section_name:
            return [f.key for f in structure['fields'] if f.section == section_name]
        return [f.key for f in structure['fields']]

    def export_structure(self, output_path: str, format: str = 'json'):
        """导出结构化文档"""
        structure = self.get_structure()

        if format == 'json':
            import json
            with open(output_path, 'w', encoding='utf-8') as f:
                # 转换dataclass为dict
                export_data = {
                    'meta': {
                        'format': structure['meta'].format.value,
                        'file_path': structure['meta'].file_path,
                        'file_size': structure['meta'].file_size,
                        'page_count': structure['meta'].page_count,
                        'total_fields': structure['meta'].total_fields
                    },
                    'sections': [
                        {
                            'title': s.title,
                            'start_row': s.start_row,
                            'end_row': s.end_row
                        } for s in structure['sections']
                    ],
                    'fields': [
                        {
                            'key': f.key,
                            'name': f.name,
                            'coord': f.coord,
                            'row': f.row,
                            'col': f.col,
                            'section': f.section
                        } for f in structure['fields']
                    ]
                }
                json.dump(export_data, f, ensure_ascii=False, indent=2)

        elif format == 'markdown':
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(structure['summary'])

    def __del__(self):
        """清理资源"""
        if self.wb:
            self.wb.close()
